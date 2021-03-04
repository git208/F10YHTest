import openpyxl

class HandleExcel:

    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_excel(self):
        """
        获取excel中的用例
        :return:
        """
        wb = openpyxl.load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        print(head_tuple)
        one_list = []
        for other_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            one_list.append(dict(zip(head_tuple, other_tuple)))
        return one_list


if __name__ == '__main__':
    das = HandleExcel('../../source/excels/接口.xlsx')

    print(das.get_excel())
