import openpyxl
import json
import time

import requests
class ParseExcel:

    def __init__(self,
                 file,
                 case_identifier=2,
                 sheet_name=None,
                 use_case_id=False
                 ):
        self.filepath = file
        self.sheetname = sheet_name
        if use_case_id == False:
            self.data = self.get_excel()[0][case_identifier-2]
        else:
            if case_identifier != None:
                self.data = self.get_excel()[1][case_identifier]
            else:
                self.data = {}

    def get_excel(self):
        """
        获取excel中的用例
        :return:
        """
        wb = openpyxl.load_workbook(self.filepath)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        case_list = []
        case_dic = {}
        for other_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            case_list.append(dict(zip(head_tuple, other_tuple)))
        for _ in case_list:
            case_dic[_['test_case_id']] = _
        return case_list,case_dic

    def data_process(self, key):
        if key in self.data.keys():
            if self.data[key] == "":
                return None
            else:
                return self.data[key]
        else:
            return None

    def get_test_case_id(self):
        return self.data_process('test_case_id')

    def get_test_case_name(self):
        return self.data_process('test_case_name')

    def get_url(self):
        return self.data_process('url')

    def get_method(self):
        return self.data_process('method')

    def get_type(self):
        return self.data_process('type')

    def get_headers(self):
        _ = self.data_process('headers')
        return json.loads(_) if _ != None else None

    def get_params(self):
        _ = self.data_process('params')
        return json.loads(_) if _ != None else None

    def get_body(self):
        _ = self.get_type().lower()
        __ = self.data_process('body')
        if _ == 'form-data' or _ == 'x-www-form-urlencoded' or _ == 'json':
            return json.loads(__) if _ != None else None
        else:
            return __


if __name__ == '__main__':
    # pass
    das = ParseExcel('../test_case/excels/接口.xlsx',case_id='GG_004')
    print(json.dumps([i for i in das.get_excel()[1].keys()],ensure_ascii=False,indent=2))
    # print(das.get_test_case_id())

