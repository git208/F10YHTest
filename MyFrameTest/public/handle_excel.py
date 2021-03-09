import openpyxl
import json
import time

import requests
class HandleExcel:

    def __init__(self, file, sheet_name=None, case_identifier=0):
        self.filepath = file
        self.sheetname = sheet_name
        self.data = self.get_excel()[case_identifier-1]

    def get_excel(self):
        """
        获取excel中的用例
        :return:
        """
        wb = openpyxl.load_workbook(self.filepath)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        print(head_tuple)
        one_list = []
        for other_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            one_list.append(dict(zip(head_tuple, other_tuple)))
        return one_list

    def data_process(self, key):
        if key in self.data.keys():
            if self.data[key] == "":
                return None
            else:
                return self.data[key]
        else:
            return None

    def get_testName(self):
        return self.data_process('title')

    def get_url(self):
        return self.data_process('url')

    def get_method(self):
        return self.data_process('method')

    def get_type(self):
        return self.data_process('type')

    def get_headers(self):
        _ = self.data_process('headers')
        return json.loads(_) if _ != None else None

    def get_params(self):
        _ = self.data_process('params')
        return json.loads(_) if _ != None else None

    def get_body(self):
        _ = self.get_type().lower()
        __ = self.data_process('body')
        if _ == 'form-data' or _ == 'x-www-form-urlencoded' or _ == 'json':
            return json.loads(__) if _ != None else None
        else:
            return __


if __name__ == '__main__':
    # pass
    das = HandleExcel('../test_case/excels/接口.xlsx', case_identifier=7)

    print(das.get_body())

